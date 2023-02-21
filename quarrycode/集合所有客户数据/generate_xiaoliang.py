#coding:utf-8
import pandas
import datetime
import pdb
from win32com.client import Dispatch, DispatchEx
from PIL import ImageGrab, Image
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import uuid
import pdb
import os

def excel_catch_custom_screen(ffile, sheetname, screen_area, img_name=False):
    ws = ffile.Sheets(sheetname)  # 选择sheet
    #pdb.set_trace()
    start=screen_area.index.start
    stop=screen_area.index.stop+2
    print(stop)
    ws.Range(ws.Cells(start,1),ws.Cells(stop,10)).CopyPicture(Format=2)  # 复制图片区域
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    imgFile = os.path.join('C:\\Users\\admin\\Documents\\yingyikuang\\snapallcustomer\\snapall\\today\\','sid'+sheetname+'-1.jpg')
    img.save(imgFile)

def excel_catch_volumn_screen(ffile,cname,len_r):
    ws = ffile.Sheets('总销量')  # 选择sheet
    ws.UsedRange.Sort(Key1=ws.Range('A1'), Order1=2)
    ws.UsedRange.AutoFilter(Field=1, Criteria1='<2021/3/22',Criteria2='>2021/3/20')
    ws.UsedRange.AutoFilter(Field=4, Criteria1=cname)
    #pdb.set_trace()
    ws.Range(ws.Cells(27610,1),ws.Cells(27748,10)).CopyPicture(Format=2)  # 复制图片区域
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    imgFile = os.path.join('C:\\Users\\admin\\Documents\\yingyikuang\\snapallcustomer\\snapall\\today\\','sid'+cname+'-2.jpg')
    img.save(imgFile)
    #pdb.set_trace()

today=datetime.date.today()
df = pandas.read_excel(u"C:\\Users\\admin\\Desktop\\客户分类.xls",sheet_name=None)
filename="C:\\Users\\admin\\Documents\\yingyikuang\\snapallcustomer\\snapall\\客户分类.xls"
filename1="C:\\Users\\admin\\Documents\\yingyikuang\\snapallcustomer\\snapall\\总销量.xlsx"
wb2=load_workbook('C:\\Users\\admin\\Documents\\yingyikuang\\snapallcustomer\\snapall\\总销量.xlsx')
tonws=wb2['总销量']
sumres=pandas.DataFrame(tonws.values)[1:]

excel = DispatchEx("Excel.Application")  # 启动excel
excel.Visible = True  # 可视化
excel.DisplayAlerts = False  # 是否显示警告
wb = excel.Workbooks.Open(filename)  # 打开excel
wb1 = excel.Workbooks.Open(filename1)  # 打开excel
date_aimday_str='2021-6-28'
date_aimday_start=datetime.datetime.strptime(date_aimday_str,'%Y-%m-%d')
date_aimday_stop=date_aimday_start+datetime.timedelta(days=1)
date_aimday=date_aimday_start.date
xres=sumres[sumres[0]>=date_aimday_start]
xres=xres[xres[0]<date_aimday_stop]
   
alist=[i for i in df]
alist.sort()
for i in alist:
    res=df[i]
    #print(i)
    #print(res)
    
    #pdb.set_trace()
    date_index=res.columns[res.isin(['DATE']).any().tolist().index(True)]
    aim_df=df[i][date_index]
    
    payment_index=res.columns[res.isin(['PAYMENT']).any().tolist().index(True)]
    balance_index=res.columns[res.isin(['BALANCE']).any().tolist().index(True)]
    amount_index=res.columns[res.isin(['AMOUNT']).any().tolist().index(True)]
    #row_date=pandas.to_datetime((row_date-25569)*86400.0,unit='s')
    #[aim_df[i] for i in range(len(aim_df)) if pandas.to_datetime((int(aim_df[i])-25569)*86400,unit='s').date()==row_date.date()]
    #找到今天，找到下一天，返回两个的行号，取2个行号之间的数值为目标数据，
    #res_aft_range=[i for i in range(len(aim_df)) if type(aim_df[i])==int and pandas.to_datetime((int(aim_df[i])-25569)*86400,unit='s').date()>date_aimday()]
    #res_bef_range=[i for i in range(len(aim_df)) if type(aim_df[i])==int and pandas.to_datetime((int(aim_df[i])-25569)*86400,unit='s').date()<date_aimday()]
    #res_curr_range=[i for i in range(len(aim_df)) if type(aim_df[i])==int and pandas.to_datetime((int(aim_df[i])-25569)*86400,unit='s').date()==date_aimday()]
    res_aft_range=[i for i in range(len(aim_df)) if isinstance(aim_df[i],datetime.datetime) and aim_df[i].date()>date_aimday()]
    res_bef_range=[i for i in range(len(aim_df)) if isinstance(aim_df[i],datetime.datetime) and aim_df[i].date()<date_aimday()]
    res_curr_range=[i for i in range(len(aim_df)) if isinstance(aim_df[i],datetime.datetime) and aim_df[i].date()==date_aimday()]
    #如果是一个数值，说明是当天，两个数值，没今天记录,是0给数值，没数据全为0
    #print(res_aft_range,res_bef_range,res_curr_range)
    if res_curr_range==[] and res_bef_range==[] and res_aft_range!=[]: #当天没记录,过去没记录，后面有记录，返回之前账户余额。过去余额为0，未来的不算，没有返利，没有账户，还是0，全返回0，
        #print(i,0,0,0,0,0)
        continue
    if res_curr_range==[] and res_bef_range==[] and res_aft_range==[]: #全没记录，返回0
        #print(i,0,0,0,0)
        continue
    if res_curr_range==[] and res_bef_range!=[] and res_aft_range==[]: #当天没记录，当天前有记录，当天后没记录，返回之前账户余额，有rebate返回rebate，没rebate，返回0
        start=res_bef_range[-1]
        stop=len(aim_df)
        bef_balance=res[balance_index][start:stop].dropna().tail(1).values[0]
        #print(i,bef_balance,0,0,bef_balance,0)
        continue
    if res_curr_range==[] and res_bef_range!=[] and res_aft_range!=[]: #当天没记录，之前有记录，之后有记录，返回之前账户余额，如果有rebate返回rebate，没rebate，返回0
        start=res_bef_range[-1]
        stop=res_aft_range[0]
        #if i=='MRAA':
        bef_balance=res[balance_index][start:stop].dropna().tail(1).values[0]
        #print(i,bef_balance,0,0,bef_balance,0)
        continue
    if res_curr_range!=[] and res_bef_range!=[] and res_aft_range==[]: #当天有记录，之前有记录,之后没记录，返回当天值和账户余额，区域直接选到frame末尾
        start=res_curr_range[0]
        stop=len(aim_df)
        ares=res[start:stop]
        bef_balance=res[balance_index][:start].dropna().tail(1).values[0]
    if res_curr_range!=[] and res_bef_range==[] and res_aft_range==[]: #当天有记录，之前没记录，之后没记录，返回当天的余额，前一天为上一行的账户余额，区域直接选到frame末尾
        start=res_curr_range[0]
        stop=len(aim_df)
        ares=res[start:stop]
        bef_balance=0
    if res_curr_range!=[] and res_bef_range!=[] and res_aft_range!=[]: #当天有记录，之前有记录，之后有记录,
        start=res_curr_range[0]
        stop=res_aft_range[0]
        bef_balance=res[balance_index][:start].dropna().tail(1).values[0]
        ares=res[start:stop]
        #if i=='KADEX':
        #    pdb.set_trace()
        balance=res[balance_index][start:stop].dropna().tail(1).values[0]
    #    print(start,stop)
    if res_curr_range!=[] and res_bef_range==[] and res_aft_range!=[]: #当天有记录，之前没有记录，之后有记录,前天余额为0，
        start=res_curr_range[0]
        stop=res_aft_range[0]
        ares=res[start:stop]
        bef_balance=0
        balance=res[balance_index][start:stop].dropna().tail(1).values[0]
    #pdb.set_trace()
    #print(ares)
    #如果某行的价格没有，跳过该行
    for item in dataframe_to_rows(ares,index=False,header=False):
        if item[2] not in [1600,3500,3700,3900,1500,3600,3800]:
            continue
        print(item[1],i,item[3],item[2],item[4])
excel.Quit()

    
