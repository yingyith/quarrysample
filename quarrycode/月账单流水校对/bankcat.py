#coding:utf-8
import pandas
import datetime
import pdb
import win32com.client
from accountmap import account_map
today=datetime.date.today()
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb2=load_workbook('C:\\Users\\admin\\Documents\\yingyikuang\\bank_month_check\\JIALONG.xlsx')
for i in wb2.sheetnames:
    aimws=wb2[i]
    rownumber=0
    res=pandas.DataFrame(aimws.values) 
    date_index=0
    delete_rows_list=[]
    for row in aimws.iter_rows(min_row=aimws.min_row,max_row=aimws.max_row,min_col=aimws.min_column,max_col=aimws.max_column,values_only=True):
        rownumber+=1
        date_check=aimws.cell(row=rownumber,column=1).value
        if date_check == None:
            print(rownumber)
            delete_rows_list.append(rownumber-1)
            continue
        if date_check in [None,'',' ','\n']==True or (isinstance(date_check,str)==True and len(date_check)!=10):
            print(rownumber)
            delete_rows_list.append(rownumber-1)
    print(delete_rows_list)
    res.drop(delete_rows_list,inplace=True)
    newres=pandas.DataFrame(columns=['start','stop','amount','balance','account'])
    prev_balance_str=res.iloc[0][7]
    print(res.iloc[0])
    print(prev_balance_str)
    prev_balance_str=''.join(prev_balance_str.split(","))
    prev_balance=float(prev_balance_str)
    for i in range(1,len(res)):
        cur_row=res.iloc[i]
        cur_row_mean=[i for i in cur_row if i!=None and isinstance(i,str)==True]
        cur_row_str=''.join(cur_row_mean)
        print(cur_row_str)
        is_have=0
        for z in account_map:
            t=account_map[z]
            for name in t:
                name_co=name.split(' ')
                print(name_co)
                check_name=[name_item in cur_row_str for name_item in name_co]
                print(check_name)
                if all(check_name)==True:
                    curr_account=z
                    is_have=1
                    break
                    #success_find add account_name as column
                else:
                    curr_account='None'
                    #fail_find  add 'not' as column 
            if is_have==1:
                break
        for j in cur_row:
            if j in [None,' ','']:
                continue
            elif isinstance(j,str)==True and len(j) >3 and len(j)<22 and len(j.split(","))>2: #and all is integer:
                curr_balance_str=''.join(j.split(","))
                print(curr_balance_str)
        #        if i==5:
        #           pdb.set_trace() 
        curr_balance=float(curr_balance_str)
        curr_amount=curr_balance-prev_balance
        prev_balance=curr_balance
        insert_row=[cur_row[0],cur_row[1],curr_amount,curr_balance,curr_account]
        print(insert_row)
        newres.loc[i-1]=insert_row

    report=newres.groupby('account').sum()
    newres.to_html('3.html')
    newres.to_csv('3.csv')
    print(report)           
    res.to_html('2.html')

