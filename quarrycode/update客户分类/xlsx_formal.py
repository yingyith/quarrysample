#coding:utf-8
import pandas
import datetime
import pdb
import win32com.client
today=datetime.date.today()
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb2=load_workbook('C:\\Users\\admin\\Documents\\yingyikuang\\test\\客户分类.xlsx')
for i in wb2.sheetnames:
    aimws=wb2[i]
    rownumber=0
    res=pandas.DataFrame(aimws.values) 
    price_index=res.columns[res.isin(['PRICE']).any().tolist().index(True)]+1
    weight_index=res.columns[res.isin(['WEIGHT']).any().tolist().index(True)]+1
    amount_index=res.columns[res.isin(['AMOUNT']).any().tolist().index(True)]+1
    payment_index=res.columns[res.isin(['PAYMENT']).any().tolist().index(True)]+1
    balance_index=res.columns[res.isin(['BALANCE']).any().tolist().index(True)]+1
    delete_condition_num=0#重复2行没重量和价格就删掉
    delete_rows_list=[]
    for row in aimws.iter_rows(min_row=aimws.min_row,max_row=aimws.max_row,min_col=aimws.min_column,max_col=aimws.max_column,values_only=True):
        rownumber+=1
        print(row)
        if aimws.cell(row=rownumber,column=price_index).value in [None,0] and aimws.cell(row=rownumber,column=weight_index).value in [None,0] and aimws.cell(row=rownumber,column=payment_index).value in [None,0]:
            delete_condition_num+=1
        else:
            delete_condition_num=0
        check_weight_cell_exist=aimws.cell(row=rownumber,column=weight_index).value
        check_price_cell_exist=aimws.cell(row=rownumber,column=price_index).value
        if isinstance(check_weight_cell_exist,str)==True and 'SUM' in check_weight_cell_exist:
            aimws.cell(row=rownumber,column=weight_index).value=0
        if isinstance(check_weight_cell_exist,str)==True and ' ' in check_weight_cell_exist:
            aimws.cell(row=rownumber,column=weight_index).value=0
        if isinstance(check_price_cell_exist,str)==True and ' ' in check_price_cell_exist:
            aimws.cell(row=rownumber,column=price_index).value=0
        if delete_condition_num>=2:
            print(rownumber)
            delete_rows_list.append(rownumber)
    print(delete_rows_list)
    for r in reversed(delete_rows_list):
        #被删行之后的行，但凡有，全要更新amount和balance列的femula
        aimws.delete_rows(r,1)
        for i in range(r,aimws.max_row+1):
    #        print(i,i-1,i-2)
            cell_row=i
            cell_amount_femula_column="="+chr(price_index-1+ord('A'))+str(cell_row)+'*'+chr(weight_index-1+ord('A'))+str(cell_row)
            aimws[i][amount_index-1].value=cell_amount_femula_column
            cell_balance_femula_column="="+chr(balance_index-1+ord('A'))+str(cell_row-1)+'+'+chr(payment_index-1+ord('A'))+str(cell_row)+'-'+chr(amount_index-1+ord('A'))+str(cell_row)
            cell_balance_femula_column_bef="="+chr(balance_index-1+ord('A'))+str(cell_row-2)+'+'+chr(payment_index-1+ord('A'))+str(cell_row-1)+'-'+chr(amount_index-1+ord('A'))+str(cell_row-1)
            aimws[i][balance_index-1].value=cell_balance_femula_column
            #aimws[i-2][balance_index-1].value=cell_balance_femula_column_bef
            
                
wb2.save('C:\\Users\\admin\\Documents\\yingyikuang\\test\\客户分类.xlsx')
office=win32com.client.Dispatch("Excel.Application")
wb4=office.Workbooks.Open('C:\\Users\\admin\\Documents\\yingyikuang\\test\\客户分类.xlsx')
wb4.RefreshAll()
wb4.Save()
wb4.Close()
