#coding:utf-8
import pandas
import datetime
import pdb
import win32com.client
today=datetime.date.today()
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb1=load_workbook('C:\\Users\\admin\\Documents\\yingyikuang\\test\\客户分类append.xlsx')
wb2=load_workbook('C:\\Users\\admin\\Documents\\yingyikuang\\test\\客户分类.xlsx')
wb3=Workbook()

for i in wb2.sheetnames:
    #if i not in ('IBB'):
    #    continue
    ws=wb1[i]
    aimws=wb2[i]
    res=pandas.DataFrame(ws.values)
    aimres=pandas.DataFrame(aimws.values)
    sheet=wb3.create_sheet(i)
    date_aimday_str='2020-12-3'
    date_aimday=datetime.datetime.strptime(date_aimday_str,'%Y-%m-%d').date()
    strp_p=datetime.datetime.strftime(date_aimday,'%m月%d日')
    date_aimday2=datetime.datetime.strptime(strp_p,'%m月%d日').date()
    date_index=res.columns[res.isin(['DATE']).any().tolist().index(True)]
    aim_df=res[date_index]
    payment_index=res.columns[res.isin(['PAYMENT']).any().tolist().index(True)]
    balance_index=res.columns[res.isin(['BALANCE']).any().tolist().index(True)]
    amount_index=res.columns[res.isin(['AMOUNT']).any().tolist().index(True)]
    price_index=res.columns[res.isin(['PRICE']).any().tolist().index(True)]
    weight_index=res.columns[res.isin(['WEIGHT']).any().tolist().index(True)]
    res_aft_range=[i for i in range(len(aim_df)) if type(aim_df[i])==int and pandas.to_datetime((int(aim_df[i])-25569)*86400,unit='s').date()>date_aimday]
    res_bef_range=[i for i in range(len(aim_df)) if type(aim_df[i])==int and pandas.to_datetime((int(aim_df[i])-25569)*86400,unit='s').date()<date_aimday]
    res_curr_range=[i for i in range(len(aim_df)) if type(aim_df[i])==int and pandas.to_datetime((int(aim_df[i])-25569)*86400,unit='s').date()==date_aimday]
    #如果是一个数值，说明是当天，两个数值，没今天记录,是0给数值，没数据全为0
    #print(res_aft_range,res_bef_range,res_curr_range)
    if res_curr_range==[] and res_bef_range==[] and res_aft_range!=[]: #当天没记录,过去没记录，后面有记录，返回之前账户余额。过去余额为0，未来的不算，没有返利，没有账户，还是0，全返回0，
       # print(i,0,0,0,0,0)
        continue
    if res_curr_range==[] and res_bef_range==[] and res_aft_range==[]: #全没记录，返回0
       # print(i,0,0,0,0)
        continue
    if res_curr_range==[] and res_bef_range!=[] and res_aft_range==[]: #当天没记录，当天前有记录，当天后没记录，返回之前账户余额，有rebate返回rebate，没rebate，返回0
        start=res_bef_range[-1]
        stop=len(aim_df)
        bef_balance=res[balance_index][start:stop].dropna().tail(1).values[0]
        #print(i,bef_balance,0,0,bef_balance,0)
        continue
    if res_curr_range==[] and res_bef_range!=[] and res_aft_range!=[]: #当天没记录，之前有记录，之后有记录，返回之前账户余额，如果有rebate返回rebate，没rebate，返回0
        start=res_bef_range[-1]
        stop=res_aft_range[0]
        bef_balance=res[balance_index][start:stop].dropna().tail(1).values[0]
        #print(i,bef_balance,0,0,bef_balance,0)
        continue
    if res_curr_range!=[] and res_bef_range!=[] and res_aft_range==[]: #当天有记录，之前有记录,之后没记录，返回当天值和账户余额，区域直接选到frame末尾
        start=res_curr_range[0]
        stop=len(aim_df)
        ares=res[start:stop]
        bef_balance=res[balance_index][:start].dropna().tail(1).values[0]
    if res_curr_range!=[] and res_bef_range==[] and res_aft_range==[]: #当天有记录，之前没记录，之后没记录，返回当天的余额，前一天为上一行的账户余额，区域直接选到frame末尾
        start=res_curr_range[0]
        stop=len(aim_df)
        ares=res[start:stop]
        bef_balance=0
    if res_curr_range!=[] and res_bef_range!=[] and res_aft_range!=[]: #当天有记录，之前有记录，之后有记录,
        start=res_curr_range[0]
        stop=res_aft_range[0]
        bef_balance=aim_df[:start].dropna().tail(1).values[0]
        balance=aim_df[start:stop].dropna().tail(1).values[0]
    #    print(start,stop)
    if res_curr_range!=[] and res_bef_range==[] and res_aft_range!=[]: #当天有记录，之前没有记录，之后有记录,前天余额为0，
        start=res_curr_range[0]
        stop=res_aft_range[0]
        bef_balance=0
        balance=aim_df[start:stop].dropna().tail(1).values[0]
    ares=res[start:stop]
    #pdb.set_trace()
    len_aimws=aimws.max_row
    ares[date_index].values[0]=date_aimday
    count=0
    print(i)
    for item in dataframe_to_rows(ares,index=False,header=False):
        cell_row=len_aimws+1+count
        cell_amount_femula_column="="+chr(price_index+ord('A'))+str(cell_row)+'*'+chr(weight_index+ord('A'))+str(cell_row)
        item[amount_index]=cell_amount_femula_column
        cell_balance_femula_column="="+chr(balance_index+ord('A'))+str(cell_row-1)+'+'+chr(payment_index+ord('A'))+str(cell_row)+'-'+chr(amount_index+ord('A'))+str(cell_row)
        item[balance_index]=cell_balance_femula_column
        aimws.append(item)
        count=count+1
    date_column=ord('A')+date_index
    aimws[chr(date_column)+str(len_aimws+1)].number_format="m月d日"
    
wb2.save('C:\\Users\\admin\\Documents\\yingyikuang\\test\\客户分类1.xlsx')
office=win32com.client.Dispatch("Excel.Application")
wb4=office.Workbooks.Open('C:\\Users\\admin\\Documents\\yingyikuang\\test\\客户分类1.xlsx')
wb4.RefreshAll()
wb4.Save()
wb4.Close()

